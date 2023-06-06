#!/usr/bin/env python
# coding: utf-8

# In[ ]:


from flask import Flask, request, jsonify, make_response, send_file
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, numbers, PatternFill
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.axis import DateAxis
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, Color, numbers
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.worksheet.hyperlink import Hyperlink
import datetime
from datetime import date
import requests
import time
import openpyxl
import calendar
import copy
import re
import boto3
import json
import io
from flask_cors import CORS

app = Flask(__name__)
CORS(app)

############################################## Start of Flask Functions ###################################################

def bank_records_types_generator(bank_data_as_dic, type_and_categories_dic):
    bank_records_array = bank_data_as_dic['Bank_Records']
    new_bank_records_array = []
    for bank_record in bank_records_array:
        record_details = bank_record['details']
        record_in = bank_record['in']
        words_dic = type_and_categories_dic['Income_dic'] if record_in != 0 else type_and_categories_dic['Outcome_dic']
        match_dic = find_bank_record_type(words_dic, record_details)
        bank_record['Type'] = match_dic['Type']
        bank_record['Category'] = match_dic['Category']
        bank_record['Sub_Category'] = match_dic['Sub_Category']
        bank_record['Match_Level'] = match_dic['Match_Level']
        bank_record['Match_Stage'] = match_dic['Match_Stage']
        new_bank_records_array.append(bank_record)
    bank_data_as_dic['Bank_Records'] = new_bank_records_array
    return bank_data_as_dic

def bank_records_daily_summary(bank_data_as_dic):
    bank_records_array = bank_data_as_dic['Bank_Records']
    daily_summary_dic = {}
    dates = [dict_['date'] for dict_ in bank_records_array if 'date' in dict_]
    if dates == sorted(dates, reverse=True):
        dates_order_from_recent_to_past = True
    else:
        dates_order_from_recent_to_past = False
    for bank_record in bank_records_array:
        record_date = bank_record['date']
        record_in = bank_record['in']
        record_out = bank_record['out']
        record_balance = bank_record['balance']
        daily_change = record_in - record_out
        before_tran_balance = (record_balance-daily_change)
        if record_date not in daily_summary_dic:
            daily_summary_dic[record_date] = {'start_of_day_balance': before_tran_balance, 'total_in': record_in, "total_out": record_out, 'daily_change': daily_change, "end_of_day_balance": record_balance}
        else:
            daily_summary_dic[record_date]['total_in'] += record_in
            daily_summary_dic[record_date]['total_out'] += record_out
            daily_summary_dic[record_date]['daily_change'] += daily_change
            if not dates_order_from_recent_to_past:
                daily_summary_dic[record_date]['start_of_day_balance'] = before_tran_balance
                daily_summary_dic[record_date]['end_of_day_balance'] = record_balance
    bank_data_as_dic['Bank_Records_Daily_Sum'] = daily_summary_dic
    return bank_data_as_dic

def generate_last_date_info(bank_data_as_dic):
    daily_summary_dic = bank_data_as_dic['Bank_Records_Daily_Sum']
    dates_of_summary = list(daily_summary_dic.keys())
    max_date = max(dates_of_summary)
    max_date_dic = copy.deepcopy(daily_summary_dic[max_date])
    max_date_dic['date'] = max_date
    bank_data_as_dic['Last_Date_Summary'] = max_date_dic
    return bank_data_as_dic
    
def find_bank_record_type(words_dic, word_to_search):
    word_to_search = word_to_search.strip()
    for TYPE in words_dic:
        identifiers = words_dic[TYPE]['Identifiers']
        for identifier in identifiers:
            if identifier in word_to_search:
                categories_names = (list(words_dic[TYPE].keys()))
                categories_names.remove("Identifiers")
                for category in categories_names:
                    sub_categories = sorted(words_dic[TYPE][category], key=len, reverse=False)
                    if word_to_search in sub_categories:
                        return {"Type": TYPE, "Category": category, "Sub_Category": word_to_search, "Match_Level": 1, "Match_Stage": 1}
                    for sub_category in sub_categories:
                        if word_to_search in sub_category:
                            match_level = len(word_to_search)/len(sub_category)
                            return {"Type": TYPE, "Category": category, "Sub_Category": sub_category, "Match_Level": match_level, "Match_Stage": 2}
    for TYPE in words_dic:
        categories_names = (list(words_dic[TYPE].keys()))
        categories_names.remove("Identifiers")
        for category in categories_names:
            sub_categories = sorted(words_dic[TYPE][category], key=len, reverse=False)
            if word_to_search in sub_categories:
                return {"Type": TYPE, "Category": category, "Sub_Category": word_to_search, "Match_Level": 1, "Match_Stage": 3}
            for sub_category in sub_categories:
                if word_to_search in sub_category:
                    match_level = len(word_to_search)/len(sub_category)
                    return {"Type": TYPE, "Category": category, "Sub_Category": sub_category, "Match_Level": match_level, "Match_Stage": 4}
    potential_matches = []
    for TYPE in words_dic:
        categories_names = (list(words_dic[TYPE].keys()))
        categories_names.remove("Identifiers")
        for category in categories_names:
            sub_categories = sorted(words_dic[TYPE][category], key=len, reverse=True)
            for sub_category in sub_categories:
                if sub_category in word_to_search:
                    potential_matches.append({"Type":TYPE, "Category":category, "Sub_Category":sub_category})
    if len(potential_matches) != 0:
        longest_sub_category_dict = {}
        longest_sub_category = ""
        max_length = 0
        for match in potential_matches:
            sub_category = match["Sub_Category"]
            if len(sub_category) > max_length:
                max_length = len(sub_category)
                longest_sub_category = sub_category
                longest_sub_category_dict = match
        match_level = len(word_to_search)/len(longest_sub_category)
        return {"Type": longest_sub_category_dict['Type'], "Category": longest_sub_category_dict['Category'], "Sub_Category": longest_sub_category, 'Match_Level': match_level, "Match_Stage": 5}
    else:
        return {"Type": "NA", "Category": "NA", "Sub_Category": "NA", 'Match_Level': "NA", "Match_Stage": "NA"}

def generate_monthly_sum(bank_data_as_dic):
    bank_records_array = bank_data_as_dic['Bank_Records']
    monthly_summary_dic = {}
    dates = [dict_['date'] for dict_ in bank_records_array if 'date' in dict_]
    if dates == sorted(dates, reverse=True):
        dates_order_from_recent_to_past = True
    else:
        dates_order_from_recent_to_past = False
    for bank_record in bank_records_array:
        if isinstance(bank_record['date'], str):
            try:
                bank_record['date'] = datetime.datetime.strptime(bank_record['date'], '%Y-%m-%d').date()
            except Exception as e:
                print(bank_record)
                print("PROBLEM")
        record_month_and_year = bank_record['date'].strftime('%m-%y')
        record_in = bank_record['in']
        record_out = bank_record['out']
        record_balance = bank_record['balance']
        daily_change = record_in - record_out
        if record_month_and_year not in monthly_summary_dic:
            monthly_summary_dic[record_month_and_year] = {'total_in': record_in, "total_out": record_out, "monthly_change": daily_change, "end_of_month_balance": record_balance}
        else:
            monthly_summary_dic[record_month_and_year]['total_in'] += record_in
            monthly_summary_dic[record_month_and_year]['total_out'] += record_out
            monthly_summary_dic[record_month_and_year]['monthly_change'] += daily_change
            if not dates_order_from_recent_to_past:
                monthly_summary_dic[record_month_and_year]['end_of_month_balance'] = record_balance
    bank_data_as_dic['Bank_Records_Monthly_Sum'] = monthly_summary_dic
    return bank_data_as_dic

def generate_types_groups(bank_data_as_dic):
    bank_records_array = bank_data_as_dic['Bank_Records']
    types_groups_dic = {'Revenues': {}, 'Expenses': {}}
    for bank_record in bank_records_array:
        record_total_in_out = bank_record['in'] + bank_record['out']
        record_type = bank_record['Type']
        record_category = bank_record['Category']
        record_sub_category = bank_record['Sub_Category']
        group_income_or_outcome = 'Revenues' if bank_record['in'] != 0 else 'Expenses'
        if record_type not in types_groups_dic[group_income_or_outcome]:
            record_type_dic = {"Total": record_total_in_out, "Categories": {}}
            record_categories_dic = {"Total": record_total_in_out, "Sub_Categories": {}}
            record_sub_categories_dic = {"Total": record_total_in_out, 'Records': [bank_record]}
            types_groups_dic[group_income_or_outcome][record_type] = record_type_dic
            types_groups_dic[group_income_or_outcome][record_type]['Categories'][record_category] = record_categories_dic
            types_groups_dic[group_income_or_outcome][record_type]['Categories'][record_category]['Sub_Categories'][record_sub_category] = record_sub_categories_dic
        elif record_category not in types_groups_dic[group_income_or_outcome][record_type]['Categories']:
            types_groups_dic[group_income_or_outcome][record_type]['Total'] += record_total_in_out
            record_categories_dic = {"Total": record_total_in_out, "Sub_Categories": {}}
            record_sub_categories_dic = {"Total": record_total_in_out, 'Records': [bank_record]}
            types_groups_dic[group_income_or_outcome][record_type]['Categories'][record_category] = record_categories_dic
            types_groups_dic[group_income_or_outcome][record_type]['Categories'][record_category]['Sub_Categories'][record_sub_category] = record_sub_categories_dic
        elif record_sub_category not in types_groups_dic[group_income_or_outcome][record_type]['Categories'][record_category]['Sub_Categories']:
            types_groups_dic[group_income_or_outcome][record_type]['Total'] += record_total_in_out
            types_groups_dic[group_income_or_outcome][record_type]['Categories'][record_category]['Total'] += record_total_in_out
            record_sub_categories_dic = {"Total": record_total_in_out, 'Records': [bank_record]}
            types_groups_dic[group_income_or_outcome][record_type]['Categories'][record_category]['Sub_Categories'][record_sub_category] = record_sub_categories_dic
        else:
            types_groups_dic[group_income_or_outcome][record_type]['Total'] += record_total_in_out
            types_groups_dic[group_income_or_outcome][record_type]['Categories'][record_category]['Total'] += record_total_in_out
            types_groups_dic[group_income_or_outcome][record_type]['Categories'][record_category]['Sub_Categories'][record_sub_category]['Total'] += record_total_in_out
            types_groups_dic[group_income_or_outcome][record_type]['Categories'][record_category]['Sub_Categories'][record_sub_category]['Records'].append(bank_record)
    bank_data_as_dic['Types_Groups_Sum'] = types_groups_dic
    return bank_data_as_dic

def generate_last_bank_records(bank_data_as_dic, number_of_records):
    bank_records_array = bank_data_as_dic['Bank_Records']
    last_bank_records = bank_records_array[-number_of_records:]
    bank_data_as_dic['Last_Bank_Records'] = last_bank_records
    return bank_data_as_dic

## Monthly Analysis Functions ---- STARTS  ###
def generate_time_range(records_array):
    records_array = sorted(records_array, key=lambda x: x['date'])
    start_date = records_array[0]['date']
    first_day_of_start_date = datetime.date(start_date.year, start_date.month, 1)
    end_date = records_array[-1]['date']
    last_day_of_end_date = end_date.replace(day=calendar.monthrange(end_date.year, end_date.month)[1])
    range_in_months = round((last_day_of_end_date - first_day_of_start_date).days/30.437)
    num_of_records = len(records_array)
    time_range_result = {'Start_Date': start_date, 'End_Date': end_date, 'Range_In_Months': range_in_months, 'Months_To_Add': None, 'Total_Range_In_Months': None, 'Total_no_Records': num_of_records, 'Average_Time_Gap': None, 'Time_Gaps': []}
    for i in range(0, len(records_array)-1):
        current_date = records_array[i]['date']
        next_date = records_array[i+1]['date']
        time_delta_in_days = (next_date - current_date).days
        time_delta_in_months = round(time_delta_in_days/30.437)
        current_time_gap_dic = {'current_date': current_date, 'next_date': next_date, 'Delta_In_Months': time_delta_in_months}
        time_range_result['Time_Gaps'].append(current_time_gap_dic)
    time_gaps_array = [record['Delta_In_Months'] for record in time_range_result['Time_Gaps']]
    unique_time_gaps = list(set(time_gaps_array))
    time_gaps_are_equal = True if len(unique_time_gaps) == 1 else False
    time_gap = unique_time_gaps[0] if time_gaps_are_equal else 0
    average_time_gap = round(sum(time_gaps_array)/len(time_gaps_array), 4) if len(time_gaps_array)!=0 else None
    time_range_result['Average_Time_Gap'] = average_time_gap
    months_to_add_to_months_range = time_gap - 1 if time_gap != 0 else 0
    time_range_result['Months_To_Add'] = months_to_add_to_months_range
    time_range_result['Total_Range_In_Months'] = range_in_months + months_to_add_to_months_range
    return time_range_result

def generate_monthly_analysis_for_records(records):
    time_range_dic = generate_time_range(records)
    total_range_in_months = time_range_dic['Total_Range_In_Months']
    total_income = sum([record['in'] for record in records])
    total_outcome = sum([record['out'] for record in records])
    monthly_average_income_for_time_range = round(total_income/total_range_in_months, 2)
    monthly_average_outcome_for_time_range = round(total_outcome/total_range_in_months, 2)
    analysis_dic = {'Total_Income':total_income, 'Avg_Monthly_Income': monthly_average_income_for_time_range, 'Total_Outcome':total_outcome, 'Avg_Monthly_Outcome': monthly_average_outcome_for_time_range, 'Time_Range_Dic': time_range_dic, 'Records': records}
    return analysis_dic

def generate_bank_account_monthly_analysis(bank_data_as_dic):
    types_groups_dic = bank_data_as_dic['Types_Groups_Sum']
    bank_account_monthly_analysis_dic = {'Revenues': {}, 'Expenses': {}}
    for revenue_or_expense in types_groups_dic:
        for TYPE in types_groups_dic[revenue_or_expense]:
            if TYPE not in bank_account_monthly_analysis_dic[revenue_or_expense]:
                bank_account_monthly_analysis_dic[revenue_or_expense][TYPE] = {}
            for category in types_groups_dic[revenue_or_expense][TYPE]['Categories']:
                category_records = []
                if category not in bank_account_monthly_analysis_dic[revenue_or_expense][TYPE]:
                    bank_account_monthly_analysis_dic[revenue_or_expense][TYPE][category] = {'Analysis_Dic': None, 'Sub_Categories': {}}
                for sub_category in types_groups_dic[revenue_or_expense][TYPE]['Categories'][category]['Sub_Categories']:
                    sub_category_records = types_groups_dic[revenue_or_expense][TYPE]['Categories'][category]['Sub_Categories'][sub_category]['Records']
                    sub_category_analysis_dic = generate_monthly_analysis_for_records(sub_category_records)
                    bank_account_monthly_analysis_dic[revenue_or_expense][TYPE][category]['Sub_Categories'][sub_category] = {'Analysis_Dic':sub_category_analysis_dic, 'Records': sub_category_records}
                    category_records += sub_category_records
                category_analysis_dic = generate_monthly_analysis_for_records(category_records)
                bank_account_monthly_analysis_dic[revenue_or_expense][TYPE][category]['Analysis_Dic'] = category_analysis_dic
    bank_data_as_dic['Monthly_Analysis'] = bank_account_monthly_analysis_dic
    return bank_data_as_dic
## Monthly Analysis Functions ---- ENDS  ###

def get_type_and_categories_dic():
    aws_access_key_id = 'AKIAX5UKROQ2H3EFZVU6'
    aws_secret_access_key = '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'
    s3 = boto3.client('s3', aws_access_key_id=aws_access_key_id,
                      aws_secret_access_key=aws_secret_access_key)
    bucket_name = 'clevirexebucket'
    object_name = 'Types.json'
    response = s3.get_object(Bucket=bucket_name, Key=object_name)
    content = response['Body'].read().decode('utf-8')
    data_dict = json.loads(content)
    return data_dict

def get_bank_account_records_analysis(bank_data_as_dic, number_of_records=10):
    type_and_categories_dic = get_type_and_categories_dic()
    bank_data_as_dic = bank_records_daily_summary(bank_data_as_dic)
    bank_data_as_dic = generate_last_date_info(bank_data_as_dic)
    bank_data_as_dic = bank_records_types_generator(bank_data_as_dic, type_and_categories_dic)
    bank_data_as_dic = generate_monthly_sum(bank_data_as_dic)
    bank_data_as_dic = generate_types_groups(bank_data_as_dic)
    bank_data_as_dic = generate_last_bank_records(bank_data_as_dic, number_of_records)
    bank_data_as_dic = generate_bank_account_monthly_analysis(bank_data_as_dic)
    # Convert all date objects to strings
    def convert_date_to_string(obj):
        if isinstance(obj, datetime.date):
            return obj.strftime('%Y-%m-%d')
        elif isinstance(obj, dict):
            return {convert_date_to_string(k): convert_date_to_string(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [convert_date_to_string(i) for i in obj]
        else:
            return obj
    bank_data_as_dic = convert_date_to_string(bank_data_as_dic)
    return bank_data_as_dic
        
def generate_bank_accounts_analysis(bank_accounts_data_array):
    ## This is each bank account analysis
    bank_accounts_analysis_array = []
    for bank_account_data_dic in bank_accounts_data_array:
        bank_accounts_analysis_array.append(get_bank_account_records_analysis(bank_account_data_dic))
    return bank_accounts_analysis_array
    
def combine_bank_accounts_analysis(bank_accounts_analysis_array):
    all_bank_accounts_analysis_dic = {'Bank_Code':'Summary', 'Account_Number':'Summary', 'Bank_Records':[], 'Bank_Records_Daily_Sum':{}, 'Last_Date_Summary':None, 'Bank_Records_Monthly_Sum':{}, 'Types_Groups_Sum':None, 'Last_Bank_Records':[], 'Monthly_Analysis':None}
    bank_accounts_analysis_array_copy = copy.deepcopy(bank_accounts_analysis_array)
    for bank_data_dic in bank_accounts_analysis_array_copy:
        current_bank_records = bank_data_dic['Bank_Records']
        all_bank_accounts_analysis_dic['Bank_Records'].extend(current_bank_records)
        ## Bank_Records_Daily_Sum Combine
        for daily_sum_dic_date in bank_data_dic['Bank_Records_Daily_Sum']:
            if daily_sum_dic_date not in all_bank_accounts_analysis_dic['Bank_Records_Daily_Sum']:
                all_bank_accounts_analysis_dic['Bank_Records_Daily_Sum'][daily_sum_dic_date] = bank_data_dic['Bank_Records_Daily_Sum'][daily_sum_dic_date]
            else:
                all_bank_accounts_analysis_dic['Bank_Records_Daily_Sum'][daily_sum_dic_date]['start_of_day_balance'] += bank_data_dic['Bank_Records_Daily_Sum'][daily_sum_dic_date]['start_of_day_balance']
                all_bank_accounts_analysis_dic['Bank_Records_Daily_Sum'][daily_sum_dic_date]['total_in'] += bank_data_dic['Bank_Records_Daily_Sum'][daily_sum_dic_date]['total_in']
                all_bank_accounts_analysis_dic['Bank_Records_Daily_Sum'][daily_sum_dic_date]['total_out'] += bank_data_dic['Bank_Records_Daily_Sum'][daily_sum_dic_date]['total_out']
                all_bank_accounts_analysis_dic['Bank_Records_Daily_Sum'][daily_sum_dic_date]['daily_change'] += bank_data_dic['Bank_Records_Daily_Sum'][daily_sum_dic_date]['daily_change']
                all_bank_accounts_analysis_dic['Bank_Records_Daily_Sum'][daily_sum_dic_date]['end_of_day_balance'] += bank_data_dic['Bank_Records_Daily_Sum'][daily_sum_dic_date]['end_of_day_balance']
        ## Last_Date_Summary Combine
        if all_bank_accounts_analysis_dic['Last_Date_Summary'] is None:
            all_bank_accounts_analysis_dic['Last_Date_Summary'] = bank_data_dic['Last_Date_Summary']
        else:
            if all_bank_accounts_analysis_dic['Last_Date_Summary']['date'] < bank_data_dic['Last_Date_Summary']['date']:
                all_bank_accounts_analysis_dic['Last_Date_Summary'] = bank_data_dic['Last_Date_Summary']
            elif all_bank_accounts_analysis_dic['Last_Date_Summary']['date'] == bank_data_dic['Last_Date_Summary']['date']:
                all_bank_accounts_analysis_dic['Last_Date_Summary']['start_of_day_balance'] += bank_data_dic['Last_Date_Summary']['start_of_day_balance']
                all_bank_accounts_analysis_dic['Last_Date_Summary']['total_in'] += bank_data_dic['Last_Date_Summary']['total_in']
                all_bank_accounts_analysis_dic['Last_Date_Summary']['total_out'] += bank_data_dic['Last_Date_Summary']['total_out']
                all_bank_accounts_analysis_dic['Last_Date_Summary']['end_of_day_balance'] += bank_data_dic['Last_Date_Summary']['end_of_day_balance']
        ## Bank_Records_Monthly_Sum Combine
        for month in bank_data_dic['Bank_Records_Monthly_Sum']:
            if month not in all_bank_accounts_analysis_dic['Bank_Records_Monthly_Sum']:
                all_bank_accounts_analysis_dic['Bank_Records_Monthly_Sum'][month] = bank_data_dic['Bank_Records_Monthly_Sum'][month]
            else:
                all_bank_accounts_analysis_dic['Bank_Records_Monthly_Sum'][month]['total_in'] += bank_data_dic['Bank_Records_Monthly_Sum'][month]['total_in']
                all_bank_accounts_analysis_dic['Bank_Records_Monthly_Sum'][month]['total_out'] += bank_data_dic['Bank_Records_Monthly_Sum'][month]['total_out']
                all_bank_accounts_analysis_dic['Bank_Records_Monthly_Sum'][month]['monthly_change'] += bank_data_dic['Bank_Records_Monthly_Sum'][month]['monthly_change']
                all_bank_accounts_analysis_dic['Bank_Records_Monthly_Sum'][month]['end_of_month_balance'] += bank_data_dic['Bank_Records_Monthly_Sum'][month]['end_of_month_balance']
        ## Types_Groups_Sum Combine
        if all_bank_accounts_analysis_dic['Types_Groups_Sum'] is None:
            all_bank_accounts_analysis_dic['Types_Groups_Sum'] = bank_data_dic['Types_Groups_Sum']
        else:
            for revenue_or_expense in bank_data_dic['Types_Groups_Sum']:
                for TYPE in bank_data_dic['Types_Groups_Sum'][revenue_or_expense]:
                    if TYPE not in all_bank_accounts_analysis_dic['Types_Groups_Sum'][revenue_or_expense]:
                        all_bank_accounts_analysis_dic['Types_Groups_Sum'][revenue_or_expense][TYPE] = bank_data_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]
                        continue
                    all_bank_accounts_analysis_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Total'] += bank_data_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Total']
                    for category in bank_data_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Categories']:
                        if category not in all_bank_accounts_analysis_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Categories']:
                            all_bank_accounts_analysis_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Categories'][category] = bank_data_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Categories'][category]
                            continue
                        all_bank_accounts_analysis_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Categories'][category]['Total'] += bank_data_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Categories'][category]['Total']
                        for sub_category in bank_data_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Categories'][category]['Sub_Categories']:
                            if sub_category not in all_bank_accounts_analysis_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Categories'][category]['Sub_Categories']:
                                all_bank_accounts_analysis_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Categories'][category]['Sub_Categories'][sub_category] = bank_data_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Categories'][category]['Sub_Categories'][sub_category]
                                continue
                            all_bank_accounts_analysis_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Categories'][category]['Sub_Categories'][sub_category]['Total'] += bank_data_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Categories'][category]['Sub_Categories'][sub_category]['Total']
                            all_bank_accounts_analysis_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Categories'][category]['Sub_Categories'][sub_category]['Records'].extend(bank_data_dic['Types_Groups_Sum'][revenue_or_expense][TYPE]['Categories'][category]['Sub_Categories'][sub_category]['Records'])
        ## Last_Bank_Records Combine
        all_bank_accounts_analysis_dic['Last_Bank_Records'] += bank_data_dic['Last_Bank_Records']
        ## Monthly_Analysis Combine
        if all_bank_accounts_analysis_dic['Monthly_Analysis'] is None:
            all_bank_accounts_analysis_dic['Monthly_Analysis'] = bank_data_dic['Monthly_Analysis']
        else:
            for revenue_or_expense in bank_data_dic['Monthly_Analysis']:
                for TYPE in bank_data_dic['Monthly_Analysis'][revenue_or_expense]:
                    if TYPE not in all_bank_accounts_analysis_dic['Monthly_Analysis'][revenue_or_expense]:
                        all_bank_accounts_analysis_dic['Monthly_Analysis'][revenue_or_expense][TYPE] = bank_data_dic['Monthly_Analysis'][revenue_or_expense][TYPE]
                        continue
                    for category in bank_data_dic['Monthly_Analysis'][revenue_or_expense][TYPE]:
                        if category not in all_bank_accounts_analysis_dic['Monthly_Analysis'][revenue_or_expense][TYPE]:
                            all_bank_accounts_analysis_dic['Monthly_Analysis'][revenue_or_expense][TYPE][category] = bank_data_dic['Monthly_Analysis'][revenue_or_expense][TYPE][category]
                            continue
                        all_bank_accounts_analysis_dic['Monthly_Analysis'][revenue_or_expense][TYPE][category]['Analysis_Dic']['Total_Income'] += bank_data_dic['Monthly_Analysis'][revenue_or_expense][TYPE][category]['Analysis_Dic']['Total_Income']
                        all_bank_accounts_analysis_dic['Monthly_Analysis'][revenue_or_expense][TYPE][category]['Analysis_Dic']['Avg_Monthly_Income'] += bank_data_dic['Monthly_Analysis'][revenue_or_expense][TYPE][category]['Analysis_Dic']['Avg_Monthly_Income']
                        all_bank_accounts_analysis_dic['Monthly_Analysis'][revenue_or_expense][TYPE][category]['Analysis_Dic']['Total_Outcome'] += bank_data_dic['Monthly_Analysis'][revenue_or_expense][TYPE][category]['Analysis_Dic']['Total_Outcome']
                        all_bank_accounts_analysis_dic['Monthly_Analysis'][revenue_or_expense][TYPE][category]['Analysis_Dic']['Avg_Monthly_Outcome'] += bank_data_dic['Monthly_Analysis'][revenue_or_expense][TYPE][category]['Analysis_Dic']['Avg_Monthly_Outcome']
                        all_bank_accounts_analysis_dic['Monthly_Analysis'][revenue_or_expense][TYPE][category]['Analysis_Dic']['Time_Range_Dic'] = None
                        all_bank_accounts_analysis_dic['Monthly_Analysis'][revenue_or_expense][TYPE][category]['Analysis_Dic']['Records'].extend(bank_data_dic['Monthly_Analysis'][revenue_or_expense][TYPE][category]['Analysis_Dic']['Records'])
    return all_bank_accounts_analysis_dic

def generate_full_bank_accounts_analysis(bank_accounts_data_array):
    bank_accounts_analysis_array = generate_bank_accounts_analysis(bank_accounts_data_array)
    all_bank_accounts_analysis_dic = combine_bank_accounts_analysis(bank_accounts_analysis_array)
    return {'Summary_Dic': all_bank_accounts_analysis_dic, 'Analysis_Array': bank_accounts_analysis_array}

######################################### GENERATE EXCEL FUNCTIONS ################################################
# Helper function to get Excel column letter based on column number
def get_excel_letter(col_num):
    return chr(ord('A') + col_num - 1)

def generate_monthly_analysis_sheet(wb, sheet_name, analysis_dic):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showGridLines = False
    ws.sheet_view.rightToLeft = True
    # Collapse the ribbon
    ws.sheet_properties.showOutlineSymbols = False
    # Define custom colors for formatting
    light_blue_fill = PatternFill(start_color='DBE5F1', end_color='DBE5F1', fill_type='solid')
    light_gray_fill = PatternFill(start_color='EDEDED', end_color='EDEDED', fill_type='solid')
    light_red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    light_green_fill = PatternFill(start_color='E5FFCC', end_color='E5FFCC', fill_type='solid')
    # Set the column widths for the headers
    start_col = 2
    ws.column_dimensions[get_excel_letter(start_col)].width = 12
    ws.column_dimensions[get_excel_letter(start_col+1)].width = 12
    ws.column_dimensions[get_excel_letter(start_col+2)].width = 12
    ws.column_dimensions[get_excel_letter(start_col+3)].width = 12
    ws.column_dimensions[get_excel_letter(start_col+4)].width = 12
    ws.column_dimensions[get_excel_letter(start_col+5)].width = 12
    ## Headers
    start_row = 2
    ws[get_excel_letter(start_col)+str(start_row)] = 'חודש'
    ws[get_excel_letter(start_col+1)+str(start_row)] = 'יתרת פתיחה'
    ws[get_excel_letter(start_col+2)+str(start_row)] = 'תקבולים'
    ws[get_excel_letter(start_col+3)+str(start_row)] = 'תשלומים'
    ws[get_excel_letter(start_col+4)+str(start_row)] = 'עודף\גרעון'
    ws[get_excel_letter(start_col+5)+str(start_row)] = 'יתרת סגירה'
    # Set the alignment for the headers
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in range(start_col, start_col+6):
        ws[get_excel_letter(col)+str(start_row)].alignment = header_alignment
    # Apply formatting to the headers
    header_font = Font(name='Arial', bold=True)
    header_border = Border(left=Side(border_style='thin', color='000000'),
                           right=Side(border_style='thin', color='000000'),
                           top=Side(border_style='thin', color='000000'),
                           bottom=Side(border_style='thin', color='000000'))
    for col in range(start_col, start_col+6):
        ws[get_excel_letter(col)+str(start_row)].fill = light_gray_fill
        ws[get_excel_letter(col)+str(start_row)].font = header_font
        ws[get_excel_letter(col)+str(start_row)].border = header_border
    row = start_row+1
    end_balance_array = []
    monthly_change_array = []
    total_out_array = []
    total_in_array = []
    start_balance_array = []
    end_balance_values_start_position = (start_col, row)
    # Write the data rows
    for month_and_year in analysis_dic['Summary_Dic']['Bank_Records_Monthly_Sum']:
        month_and_year_str = str(datetime.datetime.strptime(month_and_year, '%m-%y').strftime('%m-%Y'))
        start_balance = analysis_dic['Summary_Dic']['Bank_Records_Monthly_Sum'][month_and_year]['end_of_month_balance']-analysis_dic['Summary_Dic']['Bank_Records_Monthly_Sum'][month_and_year]['monthly_change']
        total_in = analysis_dic['Summary_Dic']['Bank_Records_Monthly_Sum'][month_and_year]['total_in']
        total_out = analysis_dic['Summary_Dic']['Bank_Records_Monthly_Sum'][month_and_year]['total_out']
        monthly_change = analysis_dic['Summary_Dic']['Bank_Records_Monthly_Sum'][month_and_year]['monthly_change']
        end_balance = analysis_dic['Summary_Dic']['Bank_Records_Monthly_Sum'][month_and_year]['end_of_month_balance']
        ws[get_excel_letter(start_col)+str(row)] = month_and_year_str
        ws[get_excel_letter(start_col+1)+str(row)] = start_balance
        ws[get_excel_letter(start_col+2)+str(row)] = total_in
        ws[get_excel_letter(start_col+3)+str(row)] = total_out
        ws[get_excel_letter(start_col+4)+str(row)] = monthly_change
        ws[get_excel_letter(start_col+5)+str(row)] = end_balance
        # Set the number format and alignment for the data cells
        data_alignment = Alignment(horizontal='right', vertical='center')
        ws[get_excel_letter(start_col)+str(row)].alignment = data_alignment
        ws[get_excel_letter(start_col+1)+str(row)].alignment = data_alignment
        ws[get_excel_letter(start_col+2)+str(row)].alignment = data_alignment
        ws[get_excel_letter(start_col+3)+str(row)].alignment = data_alignment
        ws[get_excel_letter(start_col+4)+str(row)].alignment = data_alignment
        ws[get_excel_letter(start_col+5)+str(row)].alignment = Alignment(horizontal='center', vertical='center')
        ws[get_excel_letter(start_col+1)+str(row)].number_format = '#,##0'
        ws[get_excel_letter(start_col+2)+str(row)].number_format = '#,##0'
        ws[get_excel_letter(start_col+3)+str(row)].number_format = '#,##0'
        ws[get_excel_letter(start_col+4)+str(row)].number_format = '#,##0'
        ws[get_excel_letter(start_col+5)+str(row)].number_format = '#,##0'
        # Apply conditional formatting to the monthly change cell
        if monthly_change < 0:
            ws[get_excel_letter(start_col+4)+str(row)].fill = light_red_fill
        else:
            ws[get_excel_letter(start_col+4)+str(row)].fill = light_green_fill
        # Add the data to the arrays for the summary rows
        end_balance_array.append(end_balance)
        monthly_change_array.append(monthly_change)
        total_out_array.append(total_out)
        total_in_array.append(total_in)
        start_balance_array.append(start_balance)
        row += 1
    end_balance_values_end_position = (start_col, row-1)
    summary_row = row
    avg_end_balance = round(sum(end_balance_array)/len(end_balance_array), 2)
    avg_monthly_change = round(sum(monthly_change_array)/len(monthly_change_array), 2)
    avg_total_out = round(sum(total_out_array)/len(total_out_array), 2)
    avg_total_in = round(sum(total_in_array)/len(total_in_array), 2)
    avg_start_balance = round(sum(start_balance_array)/len(start_balance_array), 2)
    # Write the summary values to the worksheet
    ws[get_excel_letter(start_col)+str(summary_row)] = 'ממוצע חודשי'
    ws[get_excel_letter(start_col+1)+str(summary_row)] = avg_start_balance
    ws[get_excel_letter(start_col+2)+str(summary_row)] = avg_total_in
    ws[get_excel_letter(start_col+3)+str(summary_row)] = avg_total_out
    ws[get_excel_letter(start_col+4)+str(summary_row)] = avg_monthly_change                                               
    ws[get_excel_letter(start_col+5)+str(summary_row)] = avg_end_balance
    
    ws[get_excel_letter(start_col+4)+str(summary_row+2)] = 'עודף \ גרעון'
    ws.merge_cells(get_excel_letter(start_col+4)+str(summary_row+2)+':'+get_excel_letter(start_col+5)+str(summary_row+2))
    ws[get_excel_letter(start_col+4)+str(summary_row+2)].font = Font(size=16, color='0000FF')
    ws[get_excel_letter(start_col+4)+str(summary_row+2)].alignment = Alignment(horizontal='center', vertical='center')
    ws[get_excel_letter(start_col+2)+str(summary_row+2)] = 'סך הכל תשלומים'
    ws.merge_cells(get_excel_letter(start_col+2)+str(summary_row+2)+':'+get_excel_letter(start_col+3)+str(summary_row+2))
    ws[get_excel_letter(start_col+2)+str(summary_row+2)].font = Font(size=16, color='0000FF')
    ws[get_excel_letter(start_col+2)+str(summary_row+2)].alignment = Alignment(horizontal='center', vertical='center')
    ws[get_excel_letter(start_col)+str(summary_row+2)] = 'סך הכל תקבולים'
    ws.merge_cells(get_excel_letter(start_col)+str(summary_row+2)+':'+get_excel_letter(start_col+1)+str(summary_row+2))
    ws[get_excel_letter(start_col)+str(summary_row+2)].font = Font(size=16, color='0000FF')
    ws[get_excel_letter(start_col)+str(summary_row+2)].alignment = Alignment(horizontal='center', vertical='center')
    ws[get_excel_letter(start_col+4)+str(summary_row+3)] = sum(monthly_change_array)
    ws.merge_cells(get_excel_letter(start_col+4)+str(summary_row+3)+':'+get_excel_letter(start_col+5)+str(summary_row+4))
    ws[get_excel_letter(start_col+4)+str(summary_row+3)].number_format = '#,##0'
    ws[get_excel_letter(start_col+4)+str(summary_row+3)].font = Font(size=24, color='000000')
    ws[get_excel_letter(start_col+4)+str(summary_row+3)].alignment = Alignment(horizontal='center', vertical='center')
    ws[get_excel_letter(start_col+2)+str(summary_row+3)] = sum(total_out_array)
    ws.merge_cells(get_excel_letter(start_col+2)+str(summary_row+3)+':'+get_excel_letter(start_col+3)+str(summary_row+4))
    ws[get_excel_letter(start_col+2)+str(summary_row+3)].number_format = '#,##0'
    ws[get_excel_letter(start_col+2)+str(summary_row+3)].font = Font(size=24, color='000000')
    ws[get_excel_letter(start_col+2)+str(summary_row+3)].alignment = Alignment(horizontal='center', vertical='center')
    ws[get_excel_letter(start_col)+str(summary_row+3)] = sum(total_in_array)
    ws.merge_cells(get_excel_letter(start_col)+str(summary_row+3)+':'+get_excel_letter(start_col+1)+str(summary_row+4))
    ws[get_excel_letter(start_col)+str(summary_row+3)].number_format = '#,##0'
    ws[get_excel_letter(start_col)+str(summary_row+3)].font = Font(size=24, color='000000')
    ws[get_excel_letter(start_col)+str(summary_row+3)].alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(border_style='thin', color='000000')
    # Define the border style
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'), 
                    diagonal=Side(style=None))
    # Define the range of cells to add the border to
    start_cell = ws[get_excel_letter(start_col)+str(summary_row+2)]
    end_cell = ws[get_excel_letter(start_col+5)+str(summary_row+4)]
    # Loop through each cell in the range
    for row in range(start_cell.row, end_cell.row + 1):
        for col in range(start_cell.column, end_cell.column + 1):
            cell = ws.cell(row=row, column=col)
            # Add border to outside cells
            if row == start_cell.row or row == end_cell.row:
                cell.border = cell.border + Border(top=thin, bottom=thin, left=thin, right=thin)
            if col == start_cell.column or col == end_cell.column:
                cell.border = cell.border + Border(top=thin, bottom=thin, left=thin, right=thin)
    ws[get_excel_letter(start_col)+str(summary_row)].alignment = data_alignment
    ws[get_excel_letter(start_col+1)+str(summary_row)].alignment = data_alignment
    ws[get_excel_letter(start_col+2)+str(summary_row)].alignment = data_alignment
    ws[get_excel_letter(start_col+3)+str(summary_row)].alignment = data_alignment
    ws[get_excel_letter(start_col+4)+str(summary_row)].alignment = data_alignment
    ws[get_excel_letter(start_col+1)+str(summary_row)].number_format = '#,##0'
    ws[get_excel_letter(start_col+2)+str(summary_row)].number_format = '#,##0'
    ws[get_excel_letter(start_col+3)+str(summary_row)].number_format = '#,##0'
    ws[get_excel_letter(start_col+4)+str(summary_row)].number_format = '#,##0'
    ws[get_excel_letter(start_col+5)+str(summary_row)].number_format = '#,##0'
    ws[get_excel_letter(start_col)+str(summary_row)].fill = light_gray_fill
    ws[get_excel_letter(start_col+1)+str(summary_row)].fill = light_gray_fill
    ws[get_excel_letter(start_col+2)+str(summary_row)].fill = light_gray_fill
    ws[get_excel_letter(start_col+3)+str(summary_row)].fill = light_gray_fill
    ws[get_excel_letter(start_col+4)+str(summary_row)].fill = light_gray_fill                                                  
                                                      
    table_range = ws[get_excel_letter(start_col)+str(start_row)+':'+get_excel_letter(start_col+5)+str(summary_row)]                                                  
    table_border = Border(left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000'))
    for row in table_range:
        for cell in row:
            cell.border = table_border                                                  
    for cell in table_range[0]:
        cell.fill = light_gray_fill
        cell.font = header_font
    for cell in table_range[-1]:
        cell.fill = light_gray_fill
        cell.font = header_font        
    if avg_monthly_change < 0:
        ws[get_excel_letter(start_col+4)+str(summary_row)].fill = light_red_fill
    else:
        ws[get_excel_letter(start_col+4)+str(summary_row)].fill = light_green_fill
    return wb

def generate_groups_analysis_sheet(wb, sheet_name, analysis_dic):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showGridLines = False
    ws.sheet_view.rightToLeft = True
    ws.title = sheet_name
    # Collapse the ribbon
    ws.sheet_properties.showOutlineSymbols = False
    # Define custom colors for formatting
    light_blue_fill = PatternFill(start_color='DBE5F1', end_color='DBE5F1', fill_type='solid')
    light_gray_fill = PatternFill(start_color='EDEDED', end_color='EDEDED', fill_type='solid')
    light_red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    light_green_fill = PatternFill(start_color='E5FFCC', end_color='E5FFCC', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(name='Arial', bold=True)
    # Set the column widths for the headers
    start_col = 2
    ws.column_dimensions[get_excel_letter(start_col)].width = 12
    ws.column_dimensions[get_excel_letter(start_col+1)].width = 12
    ws.column_dimensions[get_excel_letter(start_col+2)].width = 12
    ws.column_dimensions[get_excel_letter(start_col+3)].width = 14
    ws.column_dimensions[get_excel_letter(start_col+4)].width = 16
    ws.column_dimensions[get_excel_letter(start_col+5)].width = 14
    ## Headers
    start_row = 3
    ws.merge_cells(get_excel_letter(start_col)+str(start_row-1)+':'+get_excel_letter(start_col+5)+str(start_row-1))
    ws[get_excel_letter(start_col)+str(start_row-1)] = 'תשלומים לפי קטגוריה'
    ws[get_excel_letter(start_col)+str(start_row-1)].alignment = header_alignment
    ws[get_excel_letter(start_col)+str(start_row-1)].font = Font(name='Arial', size='14', bold=True)
    ws[get_excel_letter(start_col)+str(start_row-1)].fill = light_gray_fill
    ws[get_excel_letter(start_col)+str(start_row)] = 'סוג'
    ws[get_excel_letter(start_col+1)+str(start_row)] = 'קטגוריה'
    ws[get_excel_letter(start_col+2)+str(start_row)] = 'פירוט'
    ws[get_excel_letter(start_col+3)+str(start_row)] = 'ממוצע חודשי'
    ws[get_excel_letter(start_col+4)+str(start_row)] = 'סך הכל לתקופה'
    ws[get_excel_letter(start_col+5)+str(start_row)] = 'מעבר לתנועות'
    # Set the alignment for the headers
    for col in range(start_col, start_col+6):
        ws[get_excel_letter(col)+str(start_row)].alignment = header_alignment
    # Apply formatting to the headers
    header_border = Border(left=Side(border_style='thin', color='000000'),
                           right=Side(border_style='thin', color='000000'),
                           top=Side(border_style='thin', color='000000'),
                           bottom=Side(border_style='thin', color='000000'))
    for col in range(start_col, start_col+6):
        ws[get_excel_letter(col)+str(start_row)].fill = light_gray_fill
        ws[get_excel_letter(col)+str(start_row)].font = header_font
        ws[get_excel_letter(col)+str(start_row)].border = header_border
    expenses_by_group_dic = analysis_dic['Summary_Dic']['Monthly_Analysis']['Expenses']
    no_sub_categories_type_names = ['העברה בנקאית']
    for TYPE in expenses_by_group_dic:
        for category in expenses_by_group_dic[TYPE]:
            sub_categories = expenses_by_group_dic[TYPE][category]['Sub_Categories']
            if True: #len(sub_categories)==1 or TYPE in no_sub_categories_type_names:
                expense_name = 'NA' if len(sub_categories)==1 else 'True'
                category_analysis_dic = expenses_by_group_dic[TYPE][category]['Analysis_Dic']
                avg_monthly_outcome = category_analysis_dic['Avg_Monthly_Outcome']
                total_outcome = category_analysis_dic['Total_Outcome']
                records = category_analysis_dic['Records']
                ## write to sheet
                start_row += 1
                ws[get_excel_letter(start_col)+str(start_row)] = TYPE
                ws[get_excel_letter(start_col+1)+str(start_row)] = category
                ws[get_excel_letter(start_col+2)+str(start_row)] = ''
                ws[get_excel_letter(start_col+3)+str(start_row)] = avg_monthly_outcome
                ws[get_excel_letter(start_col+4)+str(start_row)] = total_outcome
                ws[get_excel_letter(start_col+5)+str(start_row)] = 'מעבר לתנועות'
                
                target_sheet_name = 'הוצאות'+'|'+TYPE+'|'+category
                wb = create_records_sheet(wb, target_sheet_name, records, sheet_name, False)
                hyperlink_cell = ws[get_excel_letter(start_col+5)+str(start_row)]
                hyperlink = Hyperlink(ref=hyperlink_cell.coordinate, location=f"'{target_sheet_name}'!A1")
                hyperlink_cell.hyperlink = hyperlink
                
                ws[get_excel_letter(start_col+3)+str(start_row)].number_format = '#,##0'
                ws[get_excel_letter(start_col+4)+str(start_row)].number_format = '#,##0'
                
            else: ## This is not implemented yet (TRUE ALWAYS)
                for sub_category in sub_categories:
                    sub_category_analysis_dic = sub_categories[sub_category]['Analysis_Dic']
                    expense_name = sub_category
                    avg_monthly_outcome = sub_category_analysis_dic['Avg_Monthly_Outcome']
                    total_outcome = sub_category_analysis_dic['Total_Outcome']
                    records = sub_category_analysis_dic['Records']
                    ## write to sheet
                    start_row += 1
                    ws[get_excel_letter(start_col)+str(start_row)] = TYPE
                    ws[get_excel_letter(start_col+1)+str(start_row)] = category
                    ws[get_excel_letter(start_col+2)+str(start_row)] = sub_category
                    ws[get_excel_letter(start_col+3)+str(start_row)] = avg_monthly_outcome
                    ws[get_excel_letter(start_col+4)+str(start_row)] = total_outcome
                    ws[get_excel_letter(start_col+5)+str(start_row)] = 'מעבר לתנועות'
                    ws[get_excel_letter(start_col+3)+str(start_row)].number_format = '#,##0'
                    ws[get_excel_letter(start_col+4)+str(start_row)].number_format = '#,##0'
    start_row += 3
    ws.merge_cells(get_excel_letter(start_col)+str(start_row-1)+':'+get_excel_letter(start_col+5)+str(start_row-1))
    ws[get_excel_letter(start_col)+str(start_row-1)] = 'תקבולים לפי קטגוריה'
    ws[get_excel_letter(start_col)+str(start_row-1)].alignment = header_alignment
    ws[get_excel_letter(start_col)+str(start_row-1)].font = Font(name='Arial', size='14', bold=True)
    ws[get_excel_letter(start_col)+str(start_row-1)].fill = light_gray_fill
    ws[get_excel_letter(start_col)+str(start_row)] = 'סוג'
    ws[get_excel_letter(start_col+1)+str(start_row)] = 'קטגוריה'
    ws[get_excel_letter(start_col+2)+str(start_row)] = 'פירוט'
    ws[get_excel_letter(start_col+3)+str(start_row)] = 'ממוצע חודשי'
    ws[get_excel_letter(start_col+4)+str(start_row)] = 'סך הכל לתקופה'
    ws[get_excel_letter(start_col+5)+str(start_row)] = 'מעבר לתנועות'
    # Set the alignment for the headers
    for col in range(start_col, start_col+6):
        ws[get_excel_letter(col)+str(start_row)].alignment = header_alignment
    # Apply formatting to the headers
    header_border = Border(left=Side(border_style='thin', color='000000'),
                           right=Side(border_style='thin', color='000000'),
                           top=Side(border_style='thin', color='000000'),
                           bottom=Side(border_style='thin', color='000000'))
    for col in range(start_col, start_col+6):
        ws[get_excel_letter(col)+str(start_row)].fill = light_gray_fill
        ws[get_excel_letter(col)+str(start_row)].font = header_font
        ws[get_excel_letter(col)+str(start_row)].border = header_border
    revenues_by_group_dic = analysis_dic['Summary_Dic']['Monthly_Analysis']['Revenues']
    for TYPE in revenues_by_group_dic:
        for category in revenues_by_group_dic[TYPE]:
            sub_categories = revenues_by_group_dic[TYPE][category]['Sub_Categories']
            if True: #len(sub_categories)==1 or TYPE in no_sub_categories_type_names:
                category_analysis_dic = revenues_by_group_dic[TYPE][category]['Analysis_Dic']
                avg_monthly_income = category_analysis_dic['Avg_Monthly_Income']
                total_income = category_analysis_dic['Total_Income']
                records = category_analysis_dic['Records']
                ## write to sheet
                start_row += 1
                ws[get_excel_letter(start_col)+str(start_row)] = TYPE
                ws[get_excel_letter(start_col+1)+str(start_row)] = category
                ws[get_excel_letter(start_col+2)+str(start_row)] = ''
                ws[get_excel_letter(start_col+3)+str(start_row)] = avg_monthly_income
                ws[get_excel_letter(start_col+4)+str(start_row)] = total_income
                ws[get_excel_letter(start_col+5)+str(start_row)] = 'מעבר לתנועות'
                target_sheet_name = 'הכנסות'+'|'+TYPE+'|'+category
                wb = create_records_sheet(wb, target_sheet_name, records, sheet_name, True)
                hyperlink_cell = ws[get_excel_letter(start_col+5)+str(start_row)]
                hyperlink = Hyperlink(ref=hyperlink_cell.coordinate, location=f"'{target_sheet_name}'!A1")
                hyperlink_cell.hyperlink = hyperlink
                ws[get_excel_letter(start_col+3)+str(start_row)].number_format = '#,##0'
                ws[get_excel_letter(start_col+4)+str(start_row)].number_format = '#,##0'
            else: ## This is not implemented yet (TRUE ALWAYS)
                for sub_category in sub_categories:
                    sub_category_analysis_dic = sub_categories[sub_category]['Analysis_Dic']
                    avg_monthly_income = sub_category_analysis_dic['Avg_Monthly_Income']
                    total_income = sub_category_analysis_dic['Total_Income']
                    records = sub_category_analysis_dic['Records']
                    ## write to sheet
                    start_row += 1
                    ws[get_excel_letter(start_col)+str(start_row)] = TYPE
                    ws[get_excel_letter(start_col+1)+str(start_row)] = category
                    ws[get_excel_letter(start_col+2)+str(start_row)] = sub_category
                    ws[get_excel_letter(start_col+3)+str(start_row)] = avg_monthly_income
                    ws[get_excel_letter(start_col+4)+str(start_row)] = total_income
                    ws[get_excel_letter(start_col+5)+str(start_row)] = 'מעבר לתנועות'
                    ws[get_excel_letter(start_col+3)+str(start_row)].number_format = '#,##0'
                    ws[get_excel_letter(start_col+4)+str(start_row)].number_format = '#,##0'
    return wb

def create_records_sheet(wb, sheet_name, records_array, main_sheet_name, is_revenue):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.sheet_view.showGridLines = False
    ws.sheet_view.rightToLeft = True
    # Collapse the ribbon
    ws.sheet_properties.showOutlineSymbols = False
    # Define custom colors for formatting
    light_blue_fill = PatternFill(start_color='DBE5F1', end_color='DBE5F1', fill_type='solid')
    light_gray_fill = PatternFill(start_color='EDEDED', end_color='EDEDED', fill_type='solid')
    light_red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    light_green_fill = PatternFill(start_color='E5FFCC', end_color='E5FFCC', fill_type='solid')
    # Set the column widths for the headers
    start_col = 2
    ws.column_dimensions[get_excel_letter(start_col)].width = 12
    ws.column_dimensions[get_excel_letter(start_col+1)].width = 12
    ws.column_dimensions[get_excel_letter(start_col+2)].width = 12
    ## Headers
    start_row = 2
    ws[get_excel_letter(start_col)+str(start_row)] = 'תאריך'
    ws[get_excel_letter(start_col+1)+str(start_row)] = 'פירוט'
    ws[get_excel_letter(start_col+2)+str(start_row)] = 'הכנסה' if is_revenue else 'הוצאה'
    # Set the alignment for the headers
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for col in range(start_col, start_col+3):
        ws[get_excel_letter(col)+str(start_row)].alignment = header_alignment
    # Apply formatting to the headers
    header_font = Font(name='Arial', bold=True)
    header_border = Border(left=Side(border_style='thin', color='000000'),
                           right=Side(border_style='thin', color='000000'),
                           top=Side(border_style='thin', color='000000'),
                           bottom=Side(border_style='thin', color='000000'))
    for col in range(start_col, start_col+3):
        ws[get_excel_letter(col)+str(start_row)].fill = light_gray_fill
        ws[get_excel_letter(col)+str(start_row)].font = header_font 
        ws[get_excel_letter(col)+str(start_row)].border = header_border
    target_sheet_name = main_sheet_name
    back_to_main_txt = 'חזרה לראשי'
    ws[get_excel_letter(start_col+6)+str(start_row)] = back_to_main_txt
    ws[get_excel_letter(start_col+6)+str(start_row)].font = Font(name='Arial', size='14', bold=True)
    ws[get_excel_letter(start_col+6)+str(start_row)].border = header_border
    ws.column_dimensions[get_excel_letter(start_col+6)].width = len(back_to_main_txt)+6
    ws[get_excel_letter(start_col+6)+str(start_row)].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    hyperlink_cell = ws[get_excel_letter(start_col+6)+str(start_row)]
    hyperlink = Hyperlink(ref=hyperlink_cell.coordinate, location=f"'{target_sheet_name}'!A1")
    hyperlink_cell.hyperlink = hyperlink
    max_details_len = 0
    for record in records_array:
        start_row += 1
        ws[get_excel_letter(start_col)+str(start_row)] = datetime.datetime.strptime(record['date'], '%Y-%m-%d').date()
        ws[get_excel_letter(start_col)+str(start_row)].number_format = 'dd/mm/yyyy'
        ws[get_excel_letter(start_col+1)+str(start_row)] = record['details']
        max_details_len = max_details_len if len(record['details'])<max_details_len else len(record['details'])
        ws[get_excel_letter(start_col+2)+str(start_row)] = record['in'] if record['in']!=0 else record['out']
        ws[get_excel_letter(start_col+2)+str(start_row)].number_format = '#,##0'
    ws.column_dimensions[get_excel_letter(start_col+1)].width = max_details_len
    return wb

def create_excel_analysis(analysis_dic):
    wb = Workbook()
    active_sheet = wb.active
    wb.remove(active_sheet)
    sheet_name = 'ניתוח חודשי'
    wb = generate_monthly_analysis_sheet(wb, sheet_name, analysis_dic)
    sheet_name = 'ניתוח לפי קטגוריות'
    wb = generate_groups_analysis_sheet(wb, sheet_name, analysis_dic)
    # Save the workbook to a BytesIO object
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    return excel_file

def write_files_to_s3_bucket(raw_file, processed_file, user_name):
    aws_access_key_id = 'AKIAX5UKROQ2H3EFZVU6'
    aws_secret_access_key = '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'
    s3 = boto3.client('s3', aws_access_key_id=aws_access_key_id, aws_secret_access_key=aws_secret_access_key)
    bucket_name = 'clevirexebucket'
    raw_file_name = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")+'.json'  
    processed_file_name = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")+'.xlsx'
    s3.put_object(Body=raw_file, Bucket=bucket_name, Key='Banks_API/'+user_name+'/RAW_Data/'+raw_file_name)
    s3.put_object(Body=processed_file, Bucket=bucket_name, Key='Banks_API/'+user_name+'/Processed_Data/'+processed_file_name)
    
def authenticate_user(user_name, password):
    aws_access_key_id = 'AKIAX5UKROQ2H3EFZVU6'
    aws_secret_access_key = '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'
    bucket_name = 'clevirexebucket'
    file_name = 'Banks_API/users.json'  # Update the S3 key
    s3 = boto3.client('s3', aws_access_key_id=aws_access_key_id, aws_secret_access_key=aws_secret_access_key)
    s3.download_file(bucket_name, file_name, 'users.json')  # Update the local file path
    with open('users.json', 'r') as file:  # Update the local file path
        users = json.load(file)
    for user in users:
        if user_name == user['user_name'] and password == user['password']:
            return True
    return False

def write_to_unknown_records(analysis_dic):
    unknown_income_records = []
    unknown_outcome_records = []
    for bank_analysis in analysis_dic['Analysis_Array']:
        bank_records = bank_analysis['Bank_Records']
        unknown_income_records.extend([record['details'].strip() for record in bank_records if record['in']!=0 and (record['Type']=='NA' or record['Type']=='לא מוגדר')])
        unknown_outcome_records.extend([record['details'].strip() for record in bank_records if record['out']!=0 and (record['Type']=='NA' or record['Type']=='לא מוגדר')])
    aws_access_key_id = 'AKIAX5UKROQ2H3EFZVU6'
    aws_secret_access_key = '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'
    bucket_name = 'clevirexebucket'
    file_name = 'unknown_records.json'
    s3 = boto3.client('s3', aws_access_key_id=aws_access_key_id, aws_secret_access_key=aws_secret_access_key)
    response = s3.get_object(Bucket=bucket_name, Key=file_name)
    content = response['Body'].read().decode('utf-8')
    unknown_records = json.loads(content)
    unknown_records['Income'].extend(unknown_income_records)
    unknown_records['Outcome'].extend(unknown_outcome_records)
    updated_unknown_records_json = json.dumps(unknown_records)
    s3.put_object(Body=updated_unknown_records_json, Bucket=bucket_name, Key=file_name)
    
############################################## END of Flask Functions ###################################################

##############################################    Start Endpoints    ###################################################

class CustomJSONEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, date):
            return obj.isoformat()  # Convert date object to ISO 8601 format
        return super().default(obj)

@app.route('/Bank_Accounts_Analysis_Excel', methods=['POST'])
def bank_accounts_analysis_excel():
    try:
        bank_accounts_data_array = request.json
        headers = request.headers
        print(headers)
        user_name = headers['userName']
        password = headers['password']
        if authenticate_user(user_name, password):
            json_data = json.dumps(bank_accounts_data_array, cls=CustomJSONEncoder)
            analysis_dic = generate_full_bank_accounts_analysis(bank_accounts_data_array)
            write_to_unknown_records(analysis_dic)
            analysis_excel_file = create_excel_analysis(analysis_dic)
            write_files_to_s3_bucket(json_data, analysis_excel_file, user_name)
            response = make_response(analysis_excel_file.getvalue())
            response.headers.set('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response.headers.set('Content-Disposition', 'attachment', filename='example.xlsx')
            return response
        else:
            return jsonify({"error": 'Not Allowed.'}), 401
    except Exception as e:
        print(e)
        return jsonify({"error": str(e)}), 400

@app.route('/Get_Bank_Analysis_Json', methods=['POST'])
def get_bank_analysis_json():
    try:
        bank_accounts_data_array = request.json
        analysis_dic = generate_full_bank_accounts_analysis(bank_accounts_data_array)
        write_to_unknown_records(analysis_dic)
        return jsonify(analysis_dic)
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    
##############################################    End Endpoints    ###################################################

if __name__ == '__main__':
    app.run(debug=False)

